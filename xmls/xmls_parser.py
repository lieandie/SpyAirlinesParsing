import os
import csv
from openpyxl import load_workbook

if __name__ == '__main__':
    cell_delimiter = ';'

    folder_relative_path = '../assets/YourBoardingPassDotAero/'
    output_file_path = '../assets/parsed/YourBoardingPassDotAero.xmls.csv'

    header = ['id', 'honorific', 'name', 'flight_code', 'departure', 'arrival', 'departure_code', 'arrival_code',
              'departure_date', 'departure_time', 'e_ticket', 'class', 'pnr', 'operated_by', 'seat', 'sequence',
              'card_number']
    cell_coordinates = ['A3', 'B3', 'A5', 'D5', 'H5', 'D7', 'H7', 'A9', 'C9', 'E13', 'H3', 'B13', 'E9', 'H11', 'H1',
                        'F3']

    boarding_pass_idx = 0
    file_counter = 1

    output_file = open(output_file_path, 'wb')
    output_file_writer = csv.writer(output_file, delimiter=cell_delimiter)
    output_file_writer.writerow(header)
    print 'Output files created'

    print 'Processing'
    for root, dirs, files in os.walk(folder_relative_path):
        for filename in files:
            print str(file_counter) + ' of ' + str(len(files))
            workbook = load_workbook(folder_relative_path + filename)
            for sheet_name in workbook.sheetnames:
                boarding_pass_row = [boarding_pass_idx]
                sheet = workbook[sheet_name]
                for cell_coordinate in cell_coordinates:
                    boarding_pass_row.append(sheet[cell_coordinate].value)
                output_file_writer.writerow(boarding_pass_row)
                boarding_pass_idx += 1
            file_counter += 1
